from pathlib import Path

from openpyxl import load_workbook

from tools.build_workbook import SHEET_ORDER
from tools.release_audit import run_release_audit


ROOT = Path(__file__).parents[1]
ATLAS = ROOT / "Atlas"


def test_release_audit_reports_green_release_with_documented_warnings():
    report = run_release_audit(ROOT)
    assert report["passed"] is True
    assert report["errors"] == []
    assert report["counts"]["products"] == 7
    assert report["counts"]["prompts"] == 18
    assert report["counts"]["workbook_sheets"] == len(SHEET_ORDER)
    assert any("low confidence" in warning.lower() for warning in report["warnings"])
    assert isinstance(report["generated_at"], str) and report["generated_at"]


def test_release_docs_exist_and_carry_governance_metadata():
    required = ["**Status:**", "**Version:**", "**Owner:**", "**Review date:**", "**Change history:**"]
    for relative in ["docs/RELEASE_CHECKLIST.md", "docs/VERIFICATION_REPORT.md"]:
        text = (ATLAS / relative).read_text(encoding="utf-8")
        assert all(field in text for field in required), relative


def test_readme_is_role_based_entry_point():
    text = (ATLAS / "README.md").read_text(encoding="utf-8")
    for section in [
        "## Founder path",
        "## Product designer path",
        "## Researcher path",
        "## Listing creator path",
        "## Reviewer path",
        "## Release commands",
    ]:
        assert section in text


def test_released_workbook_matches_sheet_contract():
    workbook = load_workbook(ATLAS / "excel" / "Atlas_Business_OS_v1.0.0.xlsx", data_only=False)
    assert workbook.sheetnames == SHEET_ORDER
