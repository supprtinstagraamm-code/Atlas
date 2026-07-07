from pathlib import Path
from zipfile import ZipFile

from openpyxl import load_workbook

from tools.build_workbook import SHEET_ORDER, build_workbook


ROOT = Path(__file__).parents[1]


def built_workbook(tmp_path):
    output = tmp_path / "atlas.xlsx"
    build_workbook(ROOT, output)
    return load_workbook(output, data_only=False)


def test_workbook_has_exact_sheet_order(tmp_path):
    workbook = built_workbook(tmp_path)
    assert workbook.sheetnames == SHEET_ORDER


def test_registry_sheets_have_freeze_filters_without_excel_tables(tmp_path):
    workbook = built_workbook(tmp_path)
    for name in ["Product Registry", "Collection Registry", "Evidence Register", "Atlas Scoring", "QA Gates"]:
        sheet = workbook[name]
        assert sheet.freeze_panes == "A4"
        assert sheet.auto_filter.ref is not None
        assert len(sheet.tables) == 0


def test_workbook_archive_does_not_emit_excel_table_parts(tmp_path):
    output = tmp_path / "atlas.xlsx"
    build_workbook(ROOT, output)
    with ZipFile(output) as archive:
        table_parts = [name for name in archive.namelist() if name.startswith("xl/tables/")]
        table_relationships = [name for name in archive.namelist() if name.startswith("xl/worksheets/_rels/")]
    assert table_parts == []
    assert table_relationships == []


def test_scoring_sheet_uses_formulas_and_protects_derived_cells(tmp_path):
    workbook = built_workbook(tmp_path)
    sheet = workbook["Atlas Scoring"]
    headers = {cell.value: cell.column for cell in sheet[3]}
    assert sheet.cell(4, headers["Weighted Base"]).value.startswith("=SUMPRODUCT(")
    assert sheet.cell(4, headers["Final Atlas Score"]).value.startswith("=")
    assert sheet.cell(4, headers["Decision"]).value.startswith("=IF(")
    assert sheet.cell(4, headers["Final Atlas Score"]).protection.locked is True
    assert sheet.protection.sheet is True


def test_editable_categorical_fields_have_validations(tmp_path):
    workbook = built_workbook(tmp_path)
    assert len(workbook["Product Registry"].data_validations.dataValidation) >= 3
    assert len(workbook["QA Gates"].data_validations.dataValidation) >= 1


def test_dashboard_contains_formula_driven_kpis(tmp_path):
    workbook = built_workbook(tmp_path)
    dashboard = workbook["Dashboard"]
    formulas = [cell.value for row in dashboard.iter_rows() for cell in row if isinstance(cell.value, str) and cell.value.startswith("=")]
    assert len(formulas) >= 8


def test_evidence_register_contains_plain_text_source_urls(tmp_path):
    workbook = built_workbook(tmp_path)
    values = [cell.value for row in workbook["Evidence Register"].iter_rows() for cell in row]
    assert any(isinstance(value, str) and value.startswith("https://www.etsy.com") for value in values)
    assert any(isinstance(value, str) and value.startswith("https://cults3d.com") for value in values)


def test_each_sheet_has_title_and_hidden_gridlines(tmp_path):
    workbook = built_workbook(tmp_path)
    for sheet in workbook.worksheets:
        assert isinstance(sheet["A1"].value, str) and sheet["A1"].value
        assert sheet.sheet_view.showGridLines is False
