"""Atlas v1 release audit."""

from __future__ import annotations

import json
import re
import sys
from datetime import datetime, timezone
from pathlib import Path
from zipfile import ZipFile

from openpyxl import load_workbook

if __package__ in {None, ""}:
    sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from tools.build_workbook import SHEET_ORDER
from tools.validate_atlas import calculate_atlas_score, load_contract, validate_repository


REQUIRED_PROMPTS = {
    "AI_BEHAVIOR.md",
    "BRAND_GUIDELINES.md",
    "BUSINESS_MODEL.md",
    "COLLECTION_RULES.md",
    "CONTENT_STRATEGY.md",
    "DESIGN_RULES.md",
    "MARKETPLACE_RULES.md",
    "NAMING_RULES.md",
    "PRICING_ENGINE.md",
    "PRINTING_RULES.md",
    "PRODUCT_DNA.md",
    "PRODUCT_SCORING.md",
    "PROMPT_LIBRARY.md",
    "QUALITY_CONTROL.md",
    "RENDER_RULES.md",
    "ROADMAP.md",
    "SEO_RULES.md",
    "SYSTEM_PROMPT.md",
}
METADATA_FIELDS = ["**Status:**", "**Version:**", "**Owner:**", "**Review date:**", "**Change history:**"]
UNFINISHED_PATTERN = re.compile(r"\b(?:TBD|TODO|FIXME)\b|\[insert[^\]]*\]", re.IGNORECASE)
LINK_PATTERN = re.compile(r"\[[^\]]+\]\((?!https?://|#)([^)]+(?:\.md|\.json|\.xlsx)(?:#[^)]+)?)\)")


def read_json(path: Path):
    return json.loads(path.read_text(encoding="utf-8"))


def atlas_paths(root: Path):
    atlas = root / "Atlas"
    return {
        "atlas": atlas,
        "products": sorted((atlas / "products" / "data").glob("*.json")),
        "prompts": sorted((atlas / "prompts").glob("*.md")),
        "docs": sorted(atlas.rglob("*.md")),
        "workbook": atlas / "excel" / "Atlas_Business_OS_v1.0.0.xlsx",
        "evidence": atlas / "research" / "evidence.json",
    }


def add_markdown_governance_checks(doc_paths: list[Path], errors: list[str]):
    for path in doc_paths:
        text = path.read_text(encoding="utf-8")
        missing = [field for field in METADATA_FIELDS if field not in text]
        if missing:
            errors.append(f"{path.relative_to(path.parents[1])}: missing governance metadata {missing}")
        if UNFINISHED_PATTERN.search(text):
            errors.append(f"{path.relative_to(path.parents[1])}: unfinished marker present")
        for target in LINK_PATTERN.findall(text):
            target_path = target.split("#", 1)[0]
            if not (path.parent / target_path).resolve().exists():
                errors.append(f"{path.relative_to(path.parents[1])}: broken relative link {target}")


def add_prompt_checks(prompt_paths: list[Path], errors: list[str]):
    prompt_names = {path.name for path in prompt_paths}
    missing = sorted(REQUIRED_PROMPTS - prompt_names)
    if missing:
        errors.append(f"Missing prompt files: {', '.join(missing)}")


def add_product_and_evidence_checks(root: Path, paths: dict, errors: list[str], warnings: list[str], counts: dict):
    contract = load_contract(root)
    evidence = read_json(paths["evidence"])
    evidence_ids = {record["evidence_id"] for record in evidence}
    products = [read_json(path) for path in paths["products"]]

    counts["products"] = len(products)
    counts["evidence_records"] = len(evidence)

    errors.extend(validate_repository(root))

    low_confidence = []
    for product in products:
        missing_evidence = sorted(set(product["evidence_ids"]) - evidence_ids)
        if missing_evidence:
            errors.append(f"{product['product_id']}: missing evidence references {missing_evidence}")
        expected_score = calculate_atlas_score(
            {key: detail["score"] for key, detail in product["scores"].items()},
            product["risk_penalty"],
            contract["factor_weights"],
        )
        if expected_score != product["atlas_score"]:
            errors.append(
                f"{product['product_id']}: atlas_score {product['atlas_score']} does not match calculated {expected_score}"
            )
        if product["confidence"] == "Low":
            low_confidence.append(product["product_id"])

    if low_confidence:
        warnings.append("Low confidence products remain in the launch set: " + ", ".join(low_confidence))


def add_workbook_checks(workbook_path: Path, errors: list[str], counts: dict):
    if not workbook_path.exists():
        errors.append(f"Missing workbook: {workbook_path}")
        return

    workbook = load_workbook(workbook_path, data_only=False)
    counts["workbook_sheets"] = len(workbook.sheetnames)
    if workbook.sheetnames != SHEET_ORDER:
        errors.append("Workbook sheet order does not match Atlas contract")

    for sheet_name in ["Product Registry", "Collection Registry", "Evidence Register", "Atlas Scoring", "QA Gates"]:
        if workbook[sheet_name].auto_filter.ref is None:
            errors.append(f"{sheet_name}: missing autofilter")
        if workbook[sheet_name].freeze_panes != "A4":
            errors.append(f"{sheet_name}: missing frozen header row")

    with ZipFile(workbook_path) as archive:
        table_parts = [name for name in archive.namelist() if name.startswith("xl/tables/")]
        table_relationships = []
        for name in archive.namelist():
            if not name.startswith("xl/worksheets/_rels/"):
                continue
            text = archive.read(name).decode("utf-8")
            if "relationships/table" in text or "../tables/" in text:
                table_relationships.append(name)
    if table_parts:
        errors.append("Workbook contains Excel table parts that break local Excel compatibility")
    if table_relationships:
        errors.append("Workbook contains worksheet table relationships that break local Excel compatibility")


def run_release_audit(root: Path) -> dict:
    root = Path(root)
    paths = atlas_paths(root)
    errors: list[str] = []
    warnings: list[str] = []
    counts = {
        "documents": len(paths["docs"]),
        "prompts": len(paths["prompts"]),
        "products": 0,
        "evidence_records": 0,
        "workbook_sheets": 0,
    }

    add_markdown_governance_checks(paths["docs"], errors)
    add_prompt_checks(paths["prompts"], errors)
    add_product_and_evidence_checks(root, paths, errors, warnings, counts)
    add_workbook_checks(paths["workbook"], errors, counts)

    return {
        "passed": len(errors) == 0,
        "errors": errors,
        "warnings": warnings,
        "counts": counts,
        "generated_at": datetime.now(timezone.utc).isoformat(timespec="seconds").replace("+00:00", "Z"),
    }


if __name__ == "__main__":
    repository_root = Path(__file__).resolve().parents[1]
    print(json.dumps(run_release_audit(repository_root), indent=2))
