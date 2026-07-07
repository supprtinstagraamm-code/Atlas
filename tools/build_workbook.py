"""Generate the Atlas Excel operating workbook from canonical JSON records."""

from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


SHEET_ORDER = [
    "Dashboard", "Product Registry", "Collection Registry", "Evidence Register",
    "Atlas Scoring", "QA Gates", "Marketplace Matrix", "Pricing", "SEO Keywords",
    "Lifecycle", "Roadmap", "Controlled Lists", "Instructions",
]

INK = "26332F"
MINERAL = "E8E2D8"
SAND = "C8B9A6"
SAGE = "718276"
WHITE = "FFFFFF"
PALE_GREEN = "DDE9E1"
PALE_AMBER = "F4E5BF"
PALE_RED = "F1D2CF"
PALE_BLUE = "DCE7EC"
GRID = "D6D0C7"
TITLE_FONT = "Aptos Display"
BODY_FONT = "Aptos"


def read_json(path: Path):
    return json.loads(path.read_text(encoding="utf-8"))


def load_data(root: Path):
    atlas = root / "Atlas"
    products = [read_json(path) for path in sorted((atlas / "products" / "data").glob("*.json"))]
    return {
        "contract": read_json(atlas / "schema" / "controlled-vocabularies.json"),
        "products": products,
        "collection": read_json(atlas / "collections" / "launch-collection.json"),
        "evidence": read_json(atlas / "research" / "evidence.json"),
    }


def title_sheet(sheet, title, subtitle):
    sheet.sheet_view.showGridLines = False
    sheet.merge_cells("A1:H1")
    sheet["A1"] = title
    sheet["A1"].font = Font(name=TITLE_FONT, size=20, bold=True, color=WHITE)
    sheet["A1"].fill = PatternFill("solid", fgColor=INK)
    sheet["A1"].alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 34
    sheet.merge_cells("A2:H2")
    sheet["A2"] = subtitle
    sheet["A2"].font = Font(name=BODY_FONT, size=10, color=INK, italic=True)
    sheet["A2"].fill = PatternFill("solid", fgColor=MINERAL)
    sheet["A2"].alignment = Alignment(vertical="center")
    sheet.row_dimensions[2].height = 25


def style_table(sheet, ref, name):
    sheet.auto_filter.ref = ref
    sheet.freeze_panes = "A4"
    for cell in sheet[3]:
        cell.font = Font(name=BODY_FONT, bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=SAGE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.row_dimensions[3].height = 34
    apply_row_stripes(sheet, ref)


def apply_row_stripes(sheet, ref):
    _, end = ref.split(":")
    max_row = sheet[end].row
    max_col = sheet[end].column
    stripe_fill = PatternFill("solid", fgColor=PALE_GREEN)
    for row in range(4, max_row + 1):
        if row % 2 == 0:
            for col in range(1, max_col + 1):
                sheet.cell(row, col).fill = stripe_fill


def fit_columns(sheet, minimum=11, maximum=42):
    for column_index, column_cells in enumerate(sheet.columns, 1):
        letter = get_column_letter(column_index)
        width = max(len(str(cell.value or "")) for cell in column_cells) + 2
        sheet.column_dimensions[letter].width = min(max(width, minimum), maximum)
    for row in sheet.iter_rows(min_row=4):
        for cell in row:
            cell.font = Font(name=BODY_FONT, size=10, color=INK)
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def add_list_validation(sheet, cell_range, values):
    validation = DataValidation(type="list", formula1='"' + ",".join(values) + '"', allow_blank=False)
    validation.error = "Choose a canonical Atlas value."
    validation.errorTitle = "Invalid value"
    validation.prompt = "Select from the canonical list."
    validation.promptTitle = "Atlas controlled value"
    sheet.add_data_validation(validation)
    validation.add(cell_range)


def add_whole_validation(sheet, cell_range, minimum, maximum):
    validation = DataValidation(type="decimal", operator="between", formula1=str(minimum), formula2=str(maximum), allow_blank=False)
    validation.error = f"Enter a value from {minimum} to {maximum}."
    sheet.add_data_validation(validation)
    validation.add(cell_range)


def set_number_format(sheet, cell_range, number_format):
    for row in sheet[cell_range]:
        for cell in row:
            cell.number_format = number_format


def add_dashboard(workbook, products):
    sheet = workbook["Dashboard"]
    title_sheet(sheet, "ATLAS / BUSINESS OPERATING SYSTEM", "Quiet Architectural · Accessible Premium · Evidence before recommendation")
    cards = [
        ("A4", "B4", "Products", "=COUNTA('Product Registry'!$A$4:$A$1003)"),
        ("D4", "E4", "Collections", "=COUNTA('Collection Registry'!$A$4:$A$103)"),
        ("G4", "H4", "Evidence records", "=COUNTA('Evidence Register'!$A$4:$A$1003)"),
        ("A7", "B7", "Prioritize", '=COUNTIF(\'Atlas Scoring\'!$U$4:$U$1003,"Prioritize")'),
        ("D7", "E7", "Validate", '=COUNTIF(\'Atlas Scoring\'!$U$4:$U$1003,"Validate")'),
        ("G7", "H7", "Revise / Reject", '=COUNTIF(\'Atlas Scoring\'!$U$4:$U$1003,"Revise")+COUNTIF(\'Atlas Scoring\'!$U$4:$U$1003,"Reject")'),
        ("A10", "B10", "Low confidence", '=COUNTIF(\'Product Registry\'!$G$4:$G$1003,"Low")'),
        ("D10", "E10", "QA hard fails", '=COUNTIF(\'QA Gates\'!$M$4:$M$1003,"Yes")'),
        ("G10", "H10", "Average score", '=IFERROR(AVERAGE(\'Atlas Scoring\'!$T$4:$T$1003),0)'),
    ]
    for label_cell, value_cell, label, formula in cards:
        sheet[label_cell] = label
        sheet[value_cell] = formula
        for ref in (label_cell, value_cell):
            sheet[ref].fill = PatternFill("solid", fgColor=MINERAL)
            sheet[ref].border = Border(bottom=Side(style="medium", color=SAND))
        sheet[label_cell].font = Font(name=BODY_FONT, bold=True, color=INK)
        sheet[value_cell].font = Font(name=TITLE_FONT, size=18, bold=True, color=SAGE)
        sheet[value_cell].number_format = "0.0" if label == "Average score" else "0"
    sheet["A13"] = "Launch decision"
    sheet["A13"].font = Font(name=TITLE_FONT, size=14, bold=True, color=INK)
    sheet.merge_cells("A14:H15")
    sheet["A14"] = "Aperture Ring Dish is the 70.7/100 hero candidate. The collection remains Low Confidence and requires prototype, cost, and product-specific market validation."
    sheet["A14"].alignment = Alignment(wrap_text=True, vertical="center")
    sheet["A14"].fill = PatternFill("solid", fgColor=PALE_AMBER)
    sheet["A14"].font = Font(name=BODY_FONT, size=11, color=INK)
    sheet["A17"] = "Decision bands"
    sheet["A18"] = "80–100 Prioritize | 70–79.99 Validate | 60–69.99 Revise | Below 60 Reject | Hard QA failure overrides score"
    sheet.merge_cells("A18:H18")
    sheet["A18"].alignment = Alignment(wrap_text=True)
    for col in range(1, 9):
        sheet.column_dimensions[chr(64 + col)].width = 18


def add_product_registry(workbook, products):
    sheet = workbook["Product Registry"]
    title_sheet(sheet, "Product Registry", "One canonical row per product; each product belongs to exactly one collection.")
    headers = ["Product ID", "Name", "Collection ID", "Status", "Version", "Lifecycle", "Confidence", "Owner", "Review Date", "Atlas Score", "Decision"]
    sheet.append(headers)
    for index, product in enumerate(products, 4):
        decision = "Prioritize" if product["atlas_score"] >= 80 else "Validate" if product["atlas_score"] >= 70 else "Revise" if product["atlas_score"] >= 60 else "Reject"
        sheet.append([product["product_id"], product["name"], product["collection_id"], product["status"], product["version"], product["lifecycle_stage"], product["confidence"], product["owner"], product["review_date"], f"='Atlas Scoring'!T{index}", f"='Atlas Scoring'!U{index}"])
    style_table(sheet, f"A3:K{3 + len(products)}", "ProductRegistryTable")
    add_list_validation(sheet, "D4:D1003", ["Draft", "Experimental", "Approved", "Released", "Deprecated"])
    add_list_validation(sheet, "F4:F1003", ["Idea", "Prototype", "Validated Product", "Hero Product", "Collection", "Bundle", "Premium Version", "Commercial License", "Physical Product", "Membership Exclusive", "Legacy Product", "Retired"])
    add_list_validation(sheet, "G4:G1003", ["High", "Medium", "Low"])
    set_number_format(sheet, "I4:I1003", "yyyy-mm-dd")
    fit_columns(sheet)


def add_collection_registry(workbook, collection):
    sheet = workbook["Collection Registry"]
    title_sheet(sheet, "Collection Registry", "Collection ownership, hero selection, editions, and long-term expansion.")
    headers = ["Collection ID", "Name", "Hero Product", "Supporting Products", "Mini", "XL", "Bundle", "Premium", "Commercial License", "Future Physical", "Status", "Version", "Review Date"]
    sheet.append(headers)
    sheet.append([collection["collection_id"], collection["name"], collection["hero_product_id"], ", ".join(collection["supporting_product_ids"]), collection["mini_version"], collection["xl_version"], collection["bundle_edition"], collection["premium_edition"], collection["commercial_license_edition"], collection["future_physical_product"], collection["status"], collection["version"], collection["review_date"]])
    style_table(sheet, "A3:M4", "CollectionRegistryTable")
    fit_columns(sheet, maximum=48)


def add_evidence_register(workbook, evidence):
    sheet = workbook["Evidence Register"]
    title_sheet(sheet, "Evidence Register", "Facts, proxies, assumptions, and unknowns remain traceable to dated sources.")
    headers = ["Evidence ID", "Claim", "Source URL", "Accessed", "Marketplace", "Evidence Type", "Confidence", "Notes"]
    sheet.append(headers)
    for record in evidence:
        sheet.append([record["evidence_id"], record["claim"], record["source_url"], record["accessed_at"], record["marketplace"], record["evidence_type"], record["confidence"], record["notes"]])
    style_table(sheet, f"A3:H{3 + len(evidence)}", "EvidenceRegisterTable")
    add_list_validation(sheet, "F4:F1003", ["Verified", "Proxy", "Assumption", "Unknown"])
    add_list_validation(sheet, "G4:G1003", ["High", "Medium", "Low"])
    fit_columns(sheet, maximum=60)
    sheet.column_dimensions["B"].width = 58
    sheet.column_dimensions["C"].width = 52
    sheet.column_dimensions["H"].width = 48


def add_scoring(workbook, products, contract):
    sheet = workbook["Atlas Scoring"]
    title_sheet(sheet, "Atlas Scoring", "Editable factor scores are 0–10. Formula cells are protected. Higher always means better.")
    factors = list(contract["factor_weights"])
    display = [key.replace("_", " ").title() for key in factors]
    headers = ["Product ID", "Name", *display, "Risk Penalty", "Hard Fail", "Weighted Base", "Final Atlas Score", "Decision", "Confidence"]
    sheet.append(headers)
    for row_index, product in enumerate(products, 4):
        values = [product["product_id"], product["name"]] + [product["scores"][key]["score"] for key in factors] + [product["risk_penalty"], "No"]
        sheet.append(values)
        sheet.cell(row_index, 19, f"=SUMPRODUCT(C{row_index}:P{row_index},'Controlled Lists'!$B$4:$B$17)*10")
        sheet.cell(row_index, 20, f"=MAX(0,S{row_index}-Q{row_index})")
        sheet.cell(row_index, 21, f'=IF(R{row_index}="Yes","Reject",IF(T{row_index}>=80,"Prioritize",IF(T{row_index}>=70,"Validate",IF(T{row_index}>=60,"Revise","Reject"))))')
        sheet.cell(row_index, 22, product["confidence"])
    style_table(sheet, f"A3:V{3 + len(products)}", "AtlasScoringTable")
    add_whole_validation(sheet, "C4:P1003", 0, 10)
    add_whole_validation(sheet, "Q4:Q1003", 0, 10)
    add_list_validation(sheet, "R4:R1003", ["No", "Yes"])
    add_list_validation(sheet, "V4:V1003", ["High", "Medium", "Low"])
    for row in sheet.iter_rows(min_row=4, max_row=1003, min_col=1, max_col=22):
        for cell in row:
            cell.protection = Protection(locked=cell.column in {1, 2, 19, 20, 21})
    sheet.protection.sheet = True
    sheet.protection.password = "atlas"
    set_number_format(sheet, "T4:T1003", "0.0")
    set_number_format(sheet, "S4:S1003", "0.0")
    sheet.conditional_formatting.add("T4:T1003", CellIsRule(operator="greaterThanOrEqual", formula=["80"], fill=PatternFill("solid", fgColor=PALE_GREEN)))
    sheet.conditional_formatting.add("T4:T1003", CellIsRule(operator="between", formula=["70", "79.99"], fill=PatternFill("solid", fgColor=PALE_BLUE)))
    sheet.conditional_formatting.add("T4:T1003", CellIsRule(operator="between", formula=["60", "69.99"], fill=PatternFill("solid", fgColor=PALE_AMBER)))
    sheet.conditional_formatting.add("T4:T1003", CellIsRule(operator="lessThan", formula=["60"], fill=PatternFill("solid", fgColor=PALE_RED)))
    fit_columns(sheet, maximum=24)


def add_qa(workbook, products):
    sheet = workbook["QA Gates"]
    title_sheet(sheet, "QA Gates", "Unknown hard-gate evidence blocks release. Record rationale and remediation in product files.")
    gates = ["Resin Printability", "Structural Integrity", "Support Complexity", "Commercial Viability", "SEO Opportunity", "Marketplace Fit", "Collection Fit", "Bundle Fit", "Brand Consistency", "Scalability"]
    headers = ["Product ID", "Name", *gates, "Hard Fail"]
    sheet.append(headers)
    for row_index, product in enumerate(products, 4):
        sheet.append([product["product_id"], product["name"], *(["Unknown"] * len(gates)), f'=IF(OR(C{row_index}="Fail",D{row_index}="Fail",H{row_index}="Fail"),"Yes","No")'])
    style_table(sheet, f"A3:M{3 + len(products)}", "QAGatesTable")
    add_list_validation(sheet, "C4:L1003", ["Pass", "Revise", "Fail", "Unknown"])
    sheet.conditional_formatting.add("C4:L1003", FormulaRule(formula=['C4="Fail"'], fill=PatternFill("solid", fgColor=PALE_RED)))
    fit_columns(sheet, maximum=25)


def add_marketplace_matrix(workbook, products):
    sheet = workbook["Marketplace Matrix"]
    title_sheet(sheet, "Marketplace Matrix", "Channel-specific role and presentation readiness for each launch hypothesis.")
    headers = ["Product ID", "Product", "Etsy", "Cults3D", "Gumroad", "Printables", "MyMiniFactory", "MakerWorld", "Ko-fi", "Primary Gap"]
    sheet.append(headers)
    for product in products:
        sheet.append([product["product_id"], product["name"], "Research", "Research", "Bundle later", "After print proof", "Premium later", "After print proof", "Member variant later", "Prototype and product-specific evidence"])
    style_table(sheet, f"A3:J{3 + len(products)}", "MarketplaceMatrixTable")
    add_list_validation(sheet, "C4:I1003", ["Not Fit", "Research", "Validate", "Ready", "Bundle later", "Premium later", "After print proof", "Member variant later"])
    fit_columns(sheet, maximum=30)


def add_pricing(workbook, products):
    sheet = workbook["Pricing"]
    title_sheet(sheet, "Pricing", "Inputs are assumptions until verified. Never infer another seller's profit from a visible price.")
    headers = ["Product ID", "Product", "Comparable Low", "Comparable High", "Currency", "Design Hours", "Prototype Cost", "Platform Fee %", "Provisional Price", "Evidence / Assumption", "Confidence"]
    sheet.append(headers)
    for product in products:
        sheet.append([product["product_id"], product["name"], None, None, "USD", None, None, None, None, "Unknown — collect comparable and internal cost data", "Low"])
    style_table(sheet, f"A3:K{3 + len(products)}", "PricingTable")
    add_list_validation(sheet, "K4:K1003", ["High", "Medium", "Low"])
    for column in ["C", "D", "G", "I"]:
        set_number_format(sheet, f"{column}4:{column}1003", '"$"#,##0.00')
    set_number_format(sheet, "H4:H1003", "0.0%")
    fit_columns(sheet, maximum=38)


def add_seo(workbook, products):
    sheet = workbook["SEO Keywords"]
    title_sheet(sheet, "SEO Keywords", "Keywords are hypotheses with evidence type and source; no fabricated volume.")
    headers = ["Product ID", "Marketplace", "Keyword", "Intent", "Evidence ID", "Evidence Type", "Confidence", "Status", "Notes"]
    sheet.append(headers)
    for product in products:
        sheet.append([product["product_id"], "Etsy", product["name"].lower() + " stl", "Product", product["evidence_ids"][0], "Proxy", "Low", "Research", "Validate current language before listing"])
    style_table(sheet, f"A3:I{3 + len(products)}", "SEOKeywordsTable")
    add_list_validation(sheet, "F4:F1003", ["Verified", "Proxy", "Assumption", "Unknown"])
    add_list_validation(sheet, "G4:G1003", ["High", "Medium", "Low"])
    fit_columns(sheet, maximum=38)


def add_lifecycle(workbook, products):
    sheet = workbook["Lifecycle"]
    title_sheet(sheet, "Lifecycle", "Milestones are available paths, not mandatory promotions.")
    headers = ["Product ID", "Product", "Current Stage", "Asset Status", "Version", "Entry Evidence", "Next Gate", "Owner", "Review Date"]
    sheet.append(headers)
    for product in products:
        sheet.append([product["product_id"], product["name"], product["lifecycle_stage"], product["status"], product["version"], ", ".join(product["evidence_ids"]), "Prototype and targeted research", product["owner"], product["review_date"]])
    style_table(sheet, f"A3:I{3 + len(products)}", "LifecycleTable")
    add_list_validation(sheet, "C4:C1003", ["Idea", "Prototype", "Validated Product", "Hero Product", "Collection", "Bundle", "Premium Version", "Commercial License", "Physical Product", "Membership Exclusive", "Legacy Product", "Retired"])
    add_list_validation(sheet, "D4:D1003", ["Draft", "Experimental", "Approved", "Released", "Deprecated"])
    fit_columns(sheet, maximum=38)


def add_roadmap(workbook):
    sheet = workbook["Roadmap"]
    title_sheet(sheet, "Roadmap", "Evidence and prototype gates precede release expansion.")
    headers = ["Phase", "Outcome", "Entry Criteria", "Exit Criteria", "Owner", "Status", "Review Date"]
    sheet.append(headers)
    rows = [
        ["Foundation", "Atlas v1 system", "Approved design", "Release audit passes", "System Owner", "In Progress", "2026-07-31"],
        ["Validation", "Tested hero and supporting products", "System complete", "Prototype and market evidence recorded", "Product Lead", "Not Started", "2026-10-07"],
        ["Portfolio", "Released collection and licenses", "Validated products", "Observed customer and support data", "Commercial Lead", "Not Started", "2027-01-15"],
        ["Expansion", "Physical and membership experiments", "Repeatable portfolio economics", "Go / no-go evidence documented", "Strategy Lead", "Not Started", "2027-07-01"],
    ]
    for row in rows:
        sheet.append(row)
    style_table(sheet, "A3:G7", "RoadmapTable")
    add_list_validation(sheet, "F4:F103", ["Not Started", "In Progress", "Blocked", "Complete"])
    fit_columns(sheet, maximum=42)


def add_controlled_lists(workbook, contract):
    sheet = workbook["Controlled Lists"]
    title_sheet(sheet, "Controlled Lists", "Canonical weights and vocabularies. Update JSON contract first, then regenerate this workbook.")
    sheet["A3"] = "Factor"
    sheet["B3"] = "Weight"
    for row, (factor, weight) in enumerate(contract["factor_weights"].items(), 4):
        sheet.cell(row, 1, factor)
        sheet.cell(row, 2, weight)
        sheet.cell(row, 2).number_format = "0%"
    categories = {
        "D3": ("Evidence Type", contract["evidence_types"]),
        "F3": ("Confidence", contract["confidence_levels"]),
        "H3": ("Asset Status", contract["asset_statuses"]),
        "J3": ("QA Status", contract["qa_statuses"]),
    }
    for start, (heading, values) in categories.items():
        cell = sheet[start]
        cell.value = heading
        for offset, value in enumerate(values, 1):
            sheet.cell(cell.row + offset, cell.column, value)
    for cell in sheet[3]:
        cell.font = Font(name=BODY_FONT, bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=SAGE)
    fit_columns(sheet)


def add_instructions(workbook):
    sheet = workbook["Instructions"]
    title_sheet(sheet, "Instructions", "Operate Atlas in sequence; never overwrite formula cells or turn Unknown into zero.")
    instructions = [
        ("1. Evidence", "Add a dated source record. Use Verified, Proxy, Assumption, or Unknown."),
        ("2. Product", "Create one product ID and assign exactly one collection."),
        ("3. QA", "Record every gate. Unknown hard-gate evidence blocks release."),
        ("4. Scoring", "Score all 14 factors from 0 to 10 and document rationales in the product record."),
        ("5. Risk", "Apply a 0–10 penalty with written reasons. Do not use it to hide a hard fail."),
        ("6. Decision", "Prioritize, Validate, Revise, or Reject. Hard fail always means Reject."),
        ("7. Marketplace", "Use channel-specific playbooks and keep listing claims consistent with evidence."),
        ("8. Lifecycle", "Advance only when entry evidence and approval criteria are recorded."),
        ("9. Regeneration", "Edit canonical JSON and rerun the workbook builder; do not hand-edit protected formulas."),
    ]
    sheet["A3"] = "Step"
    sheet["B3"] = "Rule"
    for row in instructions:
        sheet.append(row)
    for cell in sheet[3]:
        cell.font = Font(name=BODY_FONT, bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=SAGE)
    sheet.column_dimensions["A"].width = 24
    sheet.column_dimensions["B"].width = 95
    for row in sheet.iter_rows(min_row=4, max_col=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.font = Font(name=BODY_FONT, size=11, color=INK)
    sheet.freeze_panes = "A4"


def build_workbook(root: Path, output_path: Path) -> Path:
    root = Path(root)
    output_path = Path(output_path)
    data = load_data(root)
    workbook = Workbook()
    workbook.remove(workbook.active)
    for name in SHEET_ORDER:
        workbook.create_sheet(name)
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"

    add_dashboard(workbook, data["products"])
    add_product_registry(workbook, data["products"])
    add_collection_registry(workbook, data["collection"])
    add_evidence_register(workbook, data["evidence"])
    add_scoring(workbook, data["products"], data["contract"])
    add_qa(workbook, data["products"])
    add_marketplace_matrix(workbook, data["products"])
    add_pricing(workbook, data["products"])
    add_seo(workbook, data["products"])
    add_lifecycle(workbook, data["products"])
    add_roadmap(workbook)
    add_controlled_lists(workbook, data["contract"])
    add_instructions(workbook)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path


if __name__ == "__main__":
    repository = Path(__file__).parents[1]
    destination = repository / "Atlas" / "excel" / "Atlas_Business_OS_v1.0.0.xlsx"
    print(build_workbook(repository, destination))
